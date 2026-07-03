#!/usr/bin/env python3
"""
One-time/offline conversion: data/Artist_Song_Recommendations.xlsx
                              -> lib/artistSongRecommendations.ts

This is never run by the Next.js app or during a Vercel build — the Excel
file is only ever parsed here, on demand, by a human re-running this script
after the spreadsheet changes. The committed .ts output is the only thing
the app actually imports at runtime.

Usage: python3 scripts/convert_recommendations.py
"""
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required to run this script: pip3 install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "data" / "Artist_Song_Recommendations.xlsx"
OUT_PATH = ROOT / "lib" / "artistSongRecommendations.ts"
SHEET_NAME = "Song Recommendations"


def parse_recommended_song(cell: str) -> dict:
    """Parses a "Song Title - Artist Name" cell into {title, artist}.
    Splits on the FIRST " - " only, so song titles that themselves contain
    a hyphen (rare in this sheet, but handled safely) still parse correctly."""
    cell = cell.strip()
    if " - " not in cell:
        raise ValueError(f"Recommendation cell missing ' - ' separator: {cell!r}")
    title, artist = cell.split(" - ", 1)
    return {"title": title.strip(), "artist": artist.strip()}


def main():
    if not XLSX_PATH.exists():
        sys.exit(f"Excel file not found at {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"Sheet {SHEET_NAME!r} not found. Sheets present: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    rows = list(ws.iter_rows(values_only=True))
    header, data_rows = rows[0], rows[1:]
    expected_header = (
        "Artist", "Genre", "Song",
        "Recommended Song 1 (Same Genre)", "Recommended Song 2 (Same Genre)",
        "Recommended Song 3 (Same Genre)", "Recommended Song 4 (Same Genre)",
        "Recommended Song 5 (Same Genre)",
    )
    if tuple(header) != expected_header:
        print(f"Warning: header mismatch.\n  expected: {expected_header}\n  got:      {header}", file=sys.stderr)

    out_rows = []
    skipped = 0
    for i, row in enumerate(data_rows, start=2):
        if row is None or any(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            print(f"Skipping row {i}: contains an empty cell -> {row}", file=sys.stderr)
            skipped += 1
            continue
        artist, genre, song, *recs = row
        try:
            recommendations = [parse_recommended_song(str(r)) for r in recs]
        except ValueError as e:
            print(f"Skipping row {i}: {e}", file=sys.stderr)
            skipped += 1
            continue
        out_rows.append({
            "artist": str(artist).strip(),
            "genre": str(genre).strip(),
            "song": str(song).strip(),
            "recommendations": recommendations,
        })

    data_literal = json.dumps(out_rows, indent=2, ensure_ascii=False)

    ts = f"""// AUTO-GENERATED from data/Artist_Song_Recommendations.xlsx (sheet: "{SHEET_NAME}").
// Regenerate with: python3 scripts/convert_recommendations.py
// Do NOT hand-edit this file, and do NOT parse the .xlsx file at runtime —
// this static array is the single source of truth for genre-matched
// recommendations used by the Genre Mix Discover page. {len(out_rows)} rows
// imported from the spreadsheet.

export interface RecommendedSong {{
  title: string;
  artist: string;
}}

export interface ArtistSongRecommendationRow {{
  artist: string;
  genre: string;
  song: string;
  /** Exactly five same-genre recommendations, in spreadsheet column order. */
  recommendations: RecommendedSong[];
}}

export const ARTIST_SONG_RECOMMENDATIONS: ArtistSongRecommendationRow[] = {data_literal};
"""

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(ts, encoding="utf-8")
    print(f"Wrote {len(out_rows)} rows ({skipped} skipped) to {OUT_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
